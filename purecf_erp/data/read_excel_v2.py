import openpyxl
import json
import os

file_path = '/Users/alle/Projects/Odoo17/purecf_erp/data/Data Pengeluaran April.xlsx'

def read_excel():
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return

    wb = openpyxl.load_workbook(file_path, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        print(f"\n--- Sheet: {sheet_name} ---")
        sheet = wb[sheet_name]
        for i, row in enumerate(sheet.iter_rows(max_row=50, values_only=True)):
            if any(row):
                print(f"Row {i+1}: {row}")

if __name__ == "__main__":
    read_excel()
